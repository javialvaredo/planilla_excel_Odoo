import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class ExcelManager:
    """Clase encargada de manejar y modificar archivos Excel."""
    def __init__(self):
        self.workbook = None
        self.file_path = None
        self.modificado_path = None

    def abrir_archivo(self, file_path):
        """Carga un archivo Excel existente."""
        try:
            self.workbook = load_workbook(file_path)
            self.file_path = file_path
            base, ext = os.path.splitext(file_path)
            self.modificado_path = f"{base}_modificado{ext}"
            return True
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")
            return False

    def limpiar_columna(self, hoja, columna="E"):
        """Deja la columna indicada sin valores."""
        ws = self.workbook[hoja]
        for fila in range(1, ws.max_row + 1):
            ws[f"{columna}{fila}"].value = None



    def formatear_fechas(self, hoja, columna_fechas="A", columna_centrar="B"):
        """
        Formatea fechas en la columna indicada y centra
        también toda la columna B (u otra que elijas).
        """
        ws = self.workbook[hoja]

        # ---- Formatear y centrar columna de fechas ----
        for celda in ws[columna_fechas]:
            if isinstance(celda.value, (datetime.datetime, datetime.date)):
                celda.number_format = "DD/MM/YY"
                celda.alignment = Alignment(horizontal="center")

        # ---- Centrar también la columna B ----
        for celda in ws[columna_centrar]:
            celda.alignment = Alignment(horizontal="center")


    def modificar_columna_condicional(self, hoja):
        """
        Modifica columnas E y F según condiciones:
        - Si columna B contiene 'Despachos' → escribe en E: =valor*0.79*TC
        - Si columna G = 'USD' → en F: =valor*TC
        - Si columna G = 'EUR' → en F: =valor*EUR
        - Si columna G = 'ARS' → deja valor sin cambios
        """
        ws = self.workbook[hoja]

        for fila in range(1, ws.max_row + 1):
            celda_b = ws[f"B{fila}"]
            celda_e = ws[f"E{fila}"]
            celda_f = ws[f"F{fila}"]
            celda_g = ws[f"G{fila}"]

            if celda_f.value is None or celda_g.value is None:
                continue

            moneda = str(celda_g.value).strip().upper()
            descripcion = str(celda_b.value).strip().upper() if celda_b.value else ""

            try:
                valor = float(celda_f.value)
            except (ValueError, TypeError):
                continue

            if "DESPACHOS" in descripcion:
                celda_e.value = f"={valor}*0.79*TC"
                celda_e.number_format = '#,##0'
            else:
                if moneda == "USD":
                    celda_f.value = f"={valor}*TC"
                elif moneda == "EUR":
                    celda_f.value = f"={valor}*EUR"
                celda_f.number_format = '#,##0'

    def modificar_despachos(self, hoja):
        """
        Modifica filas donde columna B contiene:
        - 'DESPACHOS' → =valor*0.79*TC (ajustado por tipo de moneda en G)
        - 'TR_EXTERIOR' u otros → =valor*TC (ajustado por tipo de moneda en G)
        
        Si moneda = 'ARS', deja el valor sin multiplicar.
        Escribe el resultado en columna E.
        """
        ws = self.workbook[hoja]

        for fila in range(1, ws.max_row + 1):  
            celda_b = ws[f"B{fila}"]
            celda_g = ws[f"G{fila}"]  # Moneda
            celda_e = ws[f"E{fila}"]
            celda_f = ws[f"F{fila}"]

            descripcion = str(celda_b.value).strip().upper() if celda_b.value else ""
            moneda = str(celda_g.value).strip().upper() if celda_g.value else ""

            # Solo procesar si hay texto en B
            if not descripcion:
                continue

            # Intentar obtener el valor numérico en F
            try:
                valor = float(celda_f.value)
            except (ValueError, TypeError):
                continue

            # Determinar el multiplicador según moneda
            if moneda == "USD":
                tc_str = "TC"
            elif moneda == "EUR":
                tc_str = "EUR"
            elif moneda == "ARS":
                tc_str = ""  # sin tipo de cambio
            else:
                tc_str = "TC"  # por defecto usar TC

            # Aplicar lógica según tipo
            if "DESPACHOS" in descripcion:
                if moneda == "ARS":
                    celda_e.value = f"={valor}*0.79"
                else:
                    celda_e.value = f"={valor}*0.79*{tc_str}"
            elif "TR_EXTERIOR" in descripcion or descripcion:
                if moneda == "ARS":
                    celda_e.value = f"={valor}"
                else:
                    celda_e.value = f"={valor}*{tc_str}"
            else:
                continue

            celda_e.number_format = '#,##0'



    def modificar_oc_nacionales(self, hoja):
        """
        Modifica las filas de la hoja donde corresponde a OC NACIONALES.
        Aplica:
        - Si moneda = 'USD' → =valor*TC
        - Si moneda = 'EUR' → =valor*EUR (por ejemplo)
        - Si moneda = 'ARS' → deja el valor sin cambio.
        Escribe el resultado en columna E.
        """
        ws = self.workbook[hoja]

        for fila in range(1, ws.max_row + 1):
            celda_g = ws[f"G{fila}"]  # Moneda
            celda_e = ws[f"E{fila}"]
            celda_f = ws[f"F{fila}"]

            moneda = str(celda_g.value).strip().upper() if celda_g.value else ""
            if not moneda:
                continue

            try:
                valor = float(celda_f.value)
            except (ValueError, TypeError):
                continue

            if moneda == "USD":
                celda_e.value = f"={valor}*TC"
            elif moneda == "EUR":
                celda_e.value = f"={valor}*EUR"
            elif moneda == "ARS":
                celda_e.value = valor  # deja el valor tal cual
            else:
                continue  # ignora otras monedas no reconocidas

            celda_e.number_format = '#,##0'
            self.formatear_fechas(hoja, columna_fechas="A", columna_centrar="B")

   
    def guardar_modificado(self):
        """Guarda el archivo modificado con sufijo _modificado y fecha/hora."""
        # Obtener fecha y hora actuales en formato corto
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        
        # Separar nombre base y extensión del archivo original
        base, ext = os.path.splitext(self.file_path)

        # Crear nuevo nombre con sufijo y timestamp
        self.modificado_path = f"{base}_{timestamp}{ext}"

        # Guardar el archivo
        self.workbook.save(self.modificado_path)
        return self.modificado_path


    def limpiar_y_modificar(self, hoja="Hoja1"):
        """Ejecuta todo el proceso: limpiar E y modificar F según G."""
        self.limpiar_columna(hoja, "E")
        self.modificar_columna_condicional(hoja)
        self.formatear_fechas(hoja, columna_fechas="A", columna_centrar="B")
        return self.guardar_modificado()

    def aplicar_despachos(self, hoja="Hoja1"):
        """Ejecuta solo la modificación de despachos y tr_exterior."""
        self.modificar_despachos(hoja)
        self.formatear_fechas(hoja, columna_fechas="A", columna_centrar="B")
        return self.guardar_modificado()

class ExcelApp(tk.Tk):
    """Clase principal que maneja la interfaz Tkinter."""
    def __init__(self):
        super().__init__()
        self.title("Odoo Excel Modifier")
        self.geometry("500x350")

        # --- Agregar icono personalizado ---
        icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except Exception:
                pass  # En caso de error (Linux/Mac), simplemente lo omite

        self.excel_manager = ExcelManager()
        self._crear_widgets()
        
    def _crear_widgets(self):
        """Crea los elementos de la interfaz."""
        self.label_archivo = tk.Label(self, text="Ningún archivo seleccionado")
        self.label_archivo.pack(pady=10)

        # --- Botón para abrir el archivo Excel ---
        btn_abrir = tk.Button(self, text="Abrir Excel", command=self.abrir_excel)
        btn_abrir.pack(pady=5)

        # --- Botón para aplicar fórmulas a OC Clientes ---
        self.btn_modificar = tk.Button(
            self,
            text="Aplicar fórmulas OC Clientes",
            command=self.ejecutar_proceso_pedidos
        )
        self.btn_modificar.pack(pady=10)
        self.btn_modificar.config(state="disabled")

        # --- Botón para aplicar fórmulas a Despachos / Tr_Exterior ---
        self.btn_despachos = tk.Button(
            self,
            text="Aplicar fórmulas Despachos / Tr_Exterior",
            command=self.ejecutar_despachos
        )
        self.btn_despachos.pack(pady=10)
        self.btn_despachos.config(state="disabled")

        # --- Nuevo botón para aplicar fórmulas a OC Nacionales ---
        self.btn_oc_nacionales = tk.Button(
            self,
            text="Aplicar fórmulas OC Nacionales",
            command=self.ejecutar_oc_nacionales
        )
        self.btn_oc_nacionales.pack(pady=10)
        self.btn_oc_nacionales.config(state="disabled")

        # --- Botón para abrir el archivo modificado ---
        self.btn_abrir_modificado = tk.Button(
            self,
            text="Abrir Excel Modificado",
            command=self.abrir_modificado
        )
        self.btn_abrir_modificado.pack(pady=10)
        self.btn_abrir_modificado.config(state="disabled")
    

    def abrir_excel(self):
        """Permite seleccionar y abrir un archivo Excel."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos de Excel", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if file_path:
            if self.excel_manager.abrir_archivo(file_path):
                self.label_archivo.config(text=f"Archivo: {os.path.basename(file_path)}")
                
                # Habilitar botones después de cargar el archivo
                self.btn_modificar.config(state="normal")
                self.btn_despachos.config(state="normal")
                self.btn_oc_nacionales.config(state="normal")  
                
        # 🔹 Obtener automáticamente el nombre de la primera hoja
        self.primera_hoja = self.excel_manager.workbook.sheetnames[0]
        


    def ejecutar_proceso_pedidos(self):
        """Ejecuta el proceso completo: limpiar E y modificar col. F según col. G."""
        if not self.excel_manager.file_path:
            messagebox.showwarning("Atención", "Primero abra un archivo Excel.")
            return
        hoja = self.primera_hoja

        modificado_path = self.excel_manager.limpiar_y_modificar(hoja)
        messagebox.showinfo("Proceso completado", f"Archivo guardado como:\n{modificado_path}")
        self.btn_abrir_modificado.config(state="normal")

    
    def ejecutar_despachos(self):
        """Ejecuta solo la modificación de filas con 'Despachos' o 'Tr_Exterior'."""
        if not self.excel_manager.file_path:
            messagebox.showwarning("Atención", "Primero abra un archivo Excel.")
            return
        
        hoja = self.primera_hoja
        modificado_path = self.excel_manager.aplicar_despachos(hoja) #priner hoja del libro

        messagebox.showinfo(
            "Proceso completado",
            f"Archivo actualizado para Despachos / Tr_Exterior:\n{modificado_path}"
        )
        self.btn_abrir_modificado.config(state="normal")


    

    def ejecutar_oc_nacionales(self):
        """Ejecuta el proceso de modificación para OC Nacionales."""
        if not self.excel_manager.file_path:
            messagebox.showwarning("Atención", "Primero abra un archivo Excel.")
            return

        hoja = self.primera_hoja

        try:
            # Llamamos al método del ExcelManager que hace la modificación
            self.excel_manager.modificar_oc_nacionales(hoja)

            # Guardamos el archivo modificado
            ruta_modificada = self.excel_manager.guardar_modificado()

            messagebox.showinfo(
                "Éxito",
                f"Se aplicaron las fórmulas de OC Nacionales correctamente.\nArchivo guardado en:\n{ruta_modificada}"
            )

            # Habilitar el botón para abrir el Excel modificado
            self.btn_abrir_modificado.config(state="normal")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al modificar el archivo:\n{e}")

    def abrir_modificado(self):
        """Abre el archivo modificado en Excel (Windows)."""
        if self.excel_manager.modificado_path and os.path.exists(self.excel_manager.modificado_path):
            try:
                if os.name == 'nt':
                    os.startfile(self.excel_manager.modificado_path)
                else:
                    subprocess.call(['open', self.excel_manager.modificado_path])
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")
        else:
            messagebox.showwarning("Atención", "No existe el archivo modificado.")



if __name__ == "__main__":
    app = ExcelApp()
    app.mainloop()
